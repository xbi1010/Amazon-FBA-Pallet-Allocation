from __future__ import annotations

import math
import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model


# -----------------------------
# Amazon pallet constraints
# -----------------------------
PALLET_LENGTH_IN = 48.0
PALLET_WIDTH_IN = 40.0
PALLET_HEIGHT_IN = 72.0
PALLET_CAPACITY_IN3 = PALLET_LENGTH_IN * PALLET_WIDTH_IN * PALLET_HEIGHT_IN

KG_TO_LB = 2.2046226218
CM_TO_IN = 0.3937007874
VOLUME_SCALE = 100  # CP-SAT uses integers; 1 unit = 0.01 in^3


# -----------------------------
# Exact column names from your files
# -----------------------------
ORDER_ARN = "Shipment ID (ARN)"
ORDER_MODEL = "Model Number"
ORDER_UPC = "正确的UPC"
ORDER_QTY = "Quantity Confirmed(boxes)"
ORDER_PALLETS = "Pallets"
ORDER_METHOD = "打托方式"
ORDER_PO = "Order/PO Number"
ORDER_FC = "Fulfillment Center"
ORDER_SYSTEM_NO = "系统单号"

PRODUCT_SKU = "产品SKU/SKU"
PRODUCT_TITLE = "产品名称/Product Title"
PRODUCT_TITLE_EN = "产品英文名称/Product Title(EN)"
PRODUCT_MODEL = "产品型号/Product model"
PRODUCT_WEIGHT_KG = "重量/Weight"
PRODUCT_LENGTH_CM = "长/Length"
PRODUCT_WIDTH_CM = "宽/Width"
PRODUCT_HEIGHT_CM = "高/Height"
PRODUCT_UPDATED = "最后更新时间/Last Updated Time"


@dataclass
class AllocationResult:
    detail: pd.DataFrame
    pallet_summary: pd.DataFrame
    arn_summary: pd.DataFrame
    issues: pd.DataFrame
    product_warnings: pd.DataFrame


class AllocationError(Exception):
    """Business-rule error for one ARN."""


def normalize_code(value: Any) -> str:
    """Normalize UPC / ARN / SKU without losing long numeric identifiers."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    # Excel sometimes converts integer-like IDs to '12345.0'
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\s+", "", text).upper()


def normalize_model(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    text = str(value).strip().upper()
    if not text or text == "NAN":
        return ""
    return re.sub(r"[^A-Z0-9]+", "", text)


def first_nonblank(values: Iterable[Any]) -> str:
    for value in values:
        if pd.notna(value) and str(value).strip():
            return str(value).strip()
    return ""


def unique_nonblank(values: Iterable[Any]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if pd.isna(value):
            continue
        text = str(value).strip()
        if not text:
            continue
        if text not in seen:
            seen.add(text)
            result.append(text)
    return result


def read_excel_sheet_with_columns(
    source: Any,
    required_columns: set[str],
    preferred_sheet: str | None = None,
) -> tuple[pd.DataFrame, str]:
    """Find the worksheet containing all required columns."""
    sheets = pd.read_excel(source, sheet_name=None, dtype=object, engine="openpyxl")

    candidates: list[tuple[str, pd.DataFrame]] = []
    if preferred_sheet and preferred_sheet in sheets:
        candidates.append((preferred_sheet, sheets[preferred_sheet]))
    candidates.extend((name, df) for name, df in sheets.items() if name != preferred_sheet)

    for name, df in candidates:
        columns = {str(c).strip() for c in df.columns}
        if required_columns.issubset(columns):
            df = df.copy()
            df.columns = [str(c).strip() for c in df.columns]
            return df, name

    missing_text = ", ".join(sorted(required_columns))
    raise ValueError(f"找不到包含这些必要列的工作表：{missing_text}")


def load_order(order_source: Any) -> pd.DataFrame:
    required = {
        ORDER_ARN,
        ORDER_MODEL,
        ORDER_UPC,
        ORDER_QTY,
        ORDER_PALLETS,
        ORDER_METHOD,
    }
    df, _ = read_excel_sheet_with_columns(order_source, required, preferred_sheet="AFC打托")

    # Remove completely irrelevant blank rows.
    useful = (
        df[ORDER_ARN].notna()
        | df[ORDER_MODEL].notna()
        | df[ORDER_UPC].notna()
        | df[ORDER_QTY].notna()
    )
    df = df.loc[useful].copy()

    df["_row_no"] = df.index + 2  # Excel row number including header row
    df["_arn_key"] = df[ORDER_ARN].map(normalize_code)
    df["_upc_key"] = df[ORDER_UPC].map(normalize_code)
    df["_model_key"] = df[ORDER_MODEL].map(normalize_model)
    df["_qty_num"] = pd.to_numeric(df[ORDER_QTY], errors="coerce")
    df["_pallets_num"] = pd.to_numeric(df[ORDER_PALLETS], errors="coerce")

    df = df[df["_arn_key"] != ""].copy()
    if df.empty:
        raise ValueError("订单中没有有效的 Shipment ID (ARN)。")

    invalid_qty = df[
        df["_qty_num"].isna()
        | (df["_qty_num"] <= 0)
        | ((df["_qty_num"] % 1).abs() > 1e-9)
    ]
    if not invalid_qty.empty:
        rows = ", ".join(map(str, invalid_qty["_row_no"].tolist()[:20]))
        raise ValueError(f"Quantity Confirmed(boxes) 必须是正整数。问题行：{rows}")

    df["_qty_num"] = df["_qty_num"].astype(int)
    return df


def load_product_master(product_source: Any) -> tuple[pd.DataFrame, pd.DataFrame]:
    required = {
        PRODUCT_SKU,
        PRODUCT_WEIGHT_KG,
        PRODUCT_LENGTH_CM,
        PRODUCT_WIDTH_CM,
        PRODUCT_HEIGHT_CM,
    }
    df, _ = read_excel_sheet_with_columns(
        product_source,
        required,
        preferred_sheet="exportProductDataDetail",
    )

    df = df.copy()
    df["_sku_key"] = df[PRODUCT_SKU].map(normalize_code)
    df = df[df["_sku_key"] != ""].copy()

    numeric_cols = [
        PRODUCT_WEIGHT_KG,
        PRODUCT_LENGTH_CM,
        PRODUCT_WIDTH_CM,
        PRODUCT_HEIGHT_CM,
    ]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df["_updated_dt"] = pd.to_datetime(df.get(PRODUCT_UPDATED), errors="coerce")

    warnings: list[dict[str, Any]] = []
    chosen_rows: list[pd.Series] = []

    for sku_key, group in df.groupby("_sku_key", sort=False):
        valid = group.dropna(subset=numeric_cols).copy()
        if valid.empty:
            # Keep one row so matching can give a precise missing-dimensions error.
            chosen = group.iloc[-1]
            warnings.append(
                {
                    "SKU/UPC": sku_key,
                    "类型": "产品资料缺少尺寸或重量",
                    "说明": "该 SKU 存在于 Product.xlsx，但重量/长/宽/高至少有一项无法读取。",
                }
            )
            chosen_rows.append(chosen)
            continue

        unique_dims = valid[numeric_cols].round(8).drop_duplicates()
        if len(unique_dims) > 1:
            warnings.append(
                {
                    "SKU/UPC": sku_key,
                    "类型": "重复 SKU 且尺寸不一致",
                    "说明": "程序将优先使用最后更新时间最新的一条产品资料。请人工核对 Product.xlsx。",
                }
            )

        # Prefer the latest updated row; duplicates with identical dimensions are harmless.
        valid = valid.sort_values("_updated_dt", na_position="first")
        chosen_rows.append(valid.iloc[-1])

    product = pd.DataFrame(chosen_rows).reset_index(drop=True)

    # Unit conversion: kg -> lb, cm -> inch
    product["Weight_lb"] = product[PRODUCT_WEIGHT_KG] * KG_TO_LB
    product["Length_in"] = product[PRODUCT_LENGTH_CM] * CM_TO_IN
    product["Width_in"] = product[PRODUCT_WIDTH_CM] * CM_TO_IN
    product["Height_in"] = product[PRODUCT_HEIGHT_CM] * CM_TO_IN
    product["Carton_Volume_in3"] = (
        product["Length_in"] * product["Width_in"] * product["Height_in"]
    )

    warning_df = pd.DataFrame(warnings, columns=["SKU/UPC", "类型", "说明"])
    return product, warning_df


def _build_product_lookups(product: pd.DataFrame) -> tuple[dict[str, pd.Series], dict[str, pd.Series]]:
    by_sku = {row["_sku_key"]: row for _, row in product.iterrows()}

    model_candidates: dict[str, list[pd.Series]] = {}
    model_cols = [c for c in [PRODUCT_TITLE, PRODUCT_TITLE_EN, PRODUCT_MODEL] if c in product.columns]
    for _, row in product.iterrows():
        for col in model_cols:
            key = normalize_model(row.get(col))
            if key:
                model_candidates.setdefault(key, []).append(row)

    by_model: dict[str, pd.Series] = {}
    for key, rows in model_candidates.items():
        unique_skus = {r["_sku_key"] for r in rows}
        if len(unique_skus) == 1:
            by_model[key] = rows[0]

    return by_sku, by_model


def enrich_order_with_product(order: pd.DataFrame, product: pd.DataFrame) -> pd.DataFrame:
    by_sku, by_model = _build_product_lookups(product)
    enriched_rows: list[dict[str, Any]] = []

    for _, order_row in order.iterrows():
        product_row: pd.Series | None = None
        match_method = ""

        upc_key = order_row["_upc_key"]
        if upc_key and upc_key in by_sku:
            product_row = by_sku[upc_key]
            match_method = "UPC 精确匹配"
        else:
            model_key = order_row["_model_key"]
            if model_key and model_key in by_model:
                product_row = by_model[model_key]
                match_method = "Model Number 备用匹配"

        record = order_row.to_dict()
        if product_row is None:
            record.update(
                {
                    "_matched": False,
                    "_match_method": "未匹配",
                    "_product_sku_key": "",
                    "_weight_lb": math.nan,
                    "_length_in": math.nan,
                    "_width_in": math.nan,
                    "_height_in": math.nan,
                    "_carton_volume_in3": math.nan,
                }
            )
        else:
            record.update(
                {
                    "_matched": True,
                    "_match_method": match_method,
                    "_product_sku_key": product_row["_sku_key"],
                    "_weight_lb": float(product_row["Weight_lb"]),
                    "_length_in": float(product_row["Length_in"]),
                    "_width_in": float(product_row["Width_in"]),
                    "_height_in": float(product_row["Height_in"]),
                    "_carton_volume_in3": float(product_row["Carton_Volume_in3"]),
                }
            )
        enriched_rows.append(record)

    return pd.DataFrame(enriched_rows)


def normalize_method(method: str) -> str:
    value = method.strip().lower()
    if "混合" in method or "mixed" in value:
        return "mixed"
    if "single sku pallet" in value or value.startswith("single"):
        return "single"
    return "unknown"


def _validate_product_dimensions(group: pd.DataFrame, arn: str) -> None:
    unmatched = group[~group["_matched"]]
    if not unmatched.empty:
        items = []
        for _, row in unmatched.iterrows():
            items.append(
                f"Excel第{int(row['_row_no'])}行 / Model={row.get(ORDER_MODEL, '')} / UPC={row.get(ORDER_UPC, '')}"
            )
        raise AllocationError("产品资料匹配失败：" + "；".join(items))

    invalid = group[
        group["_carton_volume_in3"].isna()
        | (group["_carton_volume_in3"] <= 0)
        | group["_weight_lb"].isna()
    ]
    if not invalid.empty:
        raise AllocationError(f"ARN {arn} 的产品资料存在无效尺寸或重量。")


def _base_detail_record(
    row: pd.Series,
    arn: str,
    pallet_no: int,
    method_label: str,
    allocated_boxes: int,
) -> dict[str, Any]:
    carton_volume = float(row["_carton_volume_in3"])
    weight_lb = float(row["_weight_lb"])
    return {
        "Shipment ID (ARN)": arn,
        "Pallet No.": pallet_no,
        "打托方式": method_label,
        "Order/PO Number": row.get(ORDER_PO, ""),
        "Fulfillment Center": row.get(ORDER_FC, ""),
        "Model Number": row.get(ORDER_MODEL, ""),
        "UPC/SKU": row.get(ORDER_UPC, ""),
        "产品匹配方式": row.get("_match_method", ""),
        "Allocated Boxes": int(allocated_boxes),
        "Weight per Box (lb)": weight_lb,
        "Length (in)": float(row["_length_in"]),
        "Width (in)": float(row["_width_in"]),
        "Height (in)": float(row["_height_in"]),
        "Volume per Box (in³)": carton_volume,
        "Allocated Weight (lb)": allocated_boxes * weight_lb,
        "Allocated Volume (in³)": allocated_boxes * carton_volume,
    }


def allocate_single_sku_pallets(group: pd.DataFrame, arn: str) -> tuple[list[dict[str, Any]], int]:
    """
    For 'single sku pallet': each SKU row has its own pallet count.
    Example: 276 boxes / 10 pallets => 6 pallets x 28 + 4 pallets x 27.
    The ARN total pallet count is the sum of each SKU's Pallets value.
    """
    work = group.copy()

    missing_pallets = work[
        work["_pallets_num"].isna()
        | (work["_pallets_num"] <= 0)
        | ((work["_pallets_num"] % 1).abs() > 1e-9)
    ]
    if not missing_pallets.empty:
        rows = ", ".join(map(str, missing_pallets["_row_no"].astype(int).tolist()))
        raise AllocationError(
            f"single sku pallet 模式要求每个 SKU 行都填写正整数 Pallets。问题行：{rows}"
        )

    work["_pallets_num"] = work["_pallets_num"].astype(int)

    # If the same SKU appears more than once in the same ARN, combine quantities and pallet counts.
    aggregated: list[pd.Series] = []
    for _, sku_group in work.groupby("_product_sku_key", sort=False):
        first = sku_group.iloc[0].copy()
        first["_qty_num"] = int(sku_group["_qty_num"].sum())
        first["_pallets_num"] = int(sku_group["_pallets_num"].sum())
        aggregated.append(first)

    details: list[dict[str, Any]] = []
    pallet_no = 1

    for row in aggregated:
        qty = int(row["_qty_num"])
        pallet_count = int(row["_pallets_num"])

        if qty < pallet_count:
            raise AllocationError(
                f"SKU {row.get(ORDER_MODEL, '')} 只有 {qty} 箱，但要求 {pallet_count} 个 single-SKU 托盘，会产生空托。"
            )

        base, remainder = divmod(qty, pallet_count)
        counts = [base + 1] * remainder + [base] * (pallet_count - remainder)

        # Equal split minimizes the maximum boxes on any pallet. If this fails volume, fixed pallet count is infeasible.
        max_boxes = max(counts)
        max_volume = max_boxes * float(row["_carton_volume_in3"])
        if max_volume > PALLET_CAPACITY_IN3 + 1e-6:
            max_fit = math.floor(PALLET_CAPACITY_IN3 / float(row["_carton_volume_in3"]))
            min_pallets_by_volume = math.ceil(qty / max_fit) if max_fit > 0 else math.inf
            raise AllocationError(
                f"SKU {row.get(ORDER_MODEL, '')} 按 {pallet_count} 托平均分配后单托体积会超限。"
                f"当前最大单托 {max_boxes} 箱，约 {max_volume:,.0f} in³；"
                f"容量上限 {PALLET_CAPACITY_IN3:,.0f} in³。按体积至少需要 {min_pallets_by_volume} 托。"
            )

        for count in counts:
            details.append(
                _base_detail_record(
                    row=row,
                    arn=arn,
                    pallet_no=pallet_no,
                    method_label="single sku pallet",
                    allocated_boxes=count,
                )
            )
            pallet_no += 1

    return details, pallet_no - 1


def _solve_mixed_distribution(
    sku_rows: list[pd.Series],
    pallet_count: int,
    time_limit_seconds: float = 4.0,
) -> list[list[int]]:
    """
    Mixed-pallet integer allocation.

    Fast path:
      Each SKU is split as evenly as mathematically possible, so every pallet gets
      either floor(qty / pallets) or ceil(qty / pallets) boxes of that SKU.
      CP-SAT only decides which pallets receive the remainder boxes, while enforcing:
      - at least 2 different SKUs per pallet
      - volume <= 48*40*72 in^3
      - balanced pallet volumes

    Fallback:
      If the strict +/-1 equal split is infeasible, allow wider quantity differences
      but minimize deviation from the ideal equal split first, then balance volume.
    """
    sku_count = len(sku_rows)
    quantities = [int(row["_qty_num"]) for row in sku_rows]
    volume_units = [
        max(1, int(round(float(row["_carton_volume_in3"]) * VOLUME_SCALE)))
        for row in sku_rows
    ]
    capacity_units = int(round(PALLET_CAPACITY_IN3 * VOLUME_SCALE))

    if sku_count < 2:
        raise AllocationError("混合打托至少需要 2 个不同 SKU。")

    if sum(quantities) < 2 * pallet_count:
        raise AllocationError(
            f"混合打托要求每托至少 2 个不同 SKU，但总箱数只有 {sum(quantities)}，"
            f"不足以覆盖 {pallet_count} 托。"
        )

    presence_capacity = sum(min(qty, pallet_count) for qty in quantities)
    if presence_capacity < 2 * pallet_count:
        raise AllocationError(
            "即使把每个 SKU 尽可能分散，也无法保证每个托盘至少包含 2 个不同 SKU。"
        )

    total_volume_units = sum(q * v for q, v in zip(quantities, volume_units))
    if total_volume_units > pallet_count * capacity_units:
        total_volume = total_volume_units / VOLUME_SCALE
        raise AllocationError(
            f"总体积 {total_volume:,.0f} in³ 超过 {pallet_count} 托的总容量 "
            f"{pallet_count * PALLET_CAPACITY_IN3:,.0f} in³，固定托数下无法完成。"
        )

    for i, v in enumerate(volume_units):
        if v > capacity_units:
            raise AllocationError(
                f"SKU {sku_rows[i].get(ORDER_MODEL, '')} 单箱体积已经超过一个托盘容量。"
            )

    # ------------------------------------------------------------
    # Fast path: strict even split (each SKU differs by at most 1 box)
    # ------------------------------------------------------------
    strict = cp_model.CpModel()
    extra_vars: dict[tuple[int, int], cp_model.IntVar] = {}
    bases: list[int] = []
    remainders: list[int] = []

    for i, qty in enumerate(quantities):
        base, remainder = divmod(qty, pallet_count)
        bases.append(base)
        remainders.append(remainder)
        if remainder > 0:
            vars_for_sku = []
            for p in range(pallet_count):
                z = strict.NewBoolVar(f"extra_{i}_{p}")
                extra_vars[i, p] = z
                vars_for_sku.append(z)
            strict.Add(sum(vars_for_sku) == remainder)

    strict_volumes: list[cp_model.IntVar] = []
    for p in range(pallet_count):
        presence_terms = []
        volume_terms = []
        constant_volume = 0

        for i in range(sku_count):
            base = bases[i]
            remainder = remainders[i]
            constant_volume += base * volume_units[i]

            if base >= 1:
                presence_terms.append(1)
            elif remainder > 0:
                presence_terms.append(extra_vars[i, p])

            if remainder > 0:
                volume_terms.append(volume_units[i] * extra_vars[i, p])

        strict.Add(sum(presence_terms) >= 2)
        vol = strict.NewIntVar(0, capacity_units, f"strict_vol_{p}")
        strict.Add(vol == constant_volume + sum(volume_terms))
        strict.Add(vol <= capacity_units)
        strict_volumes.append(vol)

    # Pallets are interchangeable; ordering volumes removes a lot of symmetric search.
    for p in range(pallet_count - 1):
        strict.Add(strict_volumes[p] >= strict_volumes[p + 1])

    strict_max = strict.NewIntVar(0, capacity_units, "strict_max")
    strict_min = strict.NewIntVar(0, capacity_units, "strict_min")
    for vol in strict_volumes:
        strict.Add(strict_max >= vol)
        strict.Add(strict_min <= vol)
    strict.Minimize(strict_max - strict_min)

    strict_solver = cp_model.CpSolver()
    strict_solver.parameters.max_time_in_seconds = time_limit_seconds
    strict_solver.parameters.num_search_workers = 8
    strict_solver.parameters.random_seed = 42

    strict_status = strict_solver.Solve(strict)
    if strict_status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        result: list[list[int]] = []
        for i in range(sku_count):
            row_counts: list[int] = []
            for p in range(pallet_count):
                extra = (
                    int(strict_solver.Value(extra_vars[i, p]))
                    if (i, p) in extra_vars
                    else 0
                )
                row_counts.append(bases[i] + extra)
            result.append(row_counts)
        return result

    # ------------------------------------------------------------
    # Fallback: relaxed equalization, still honoring all hard rules
    # ------------------------------------------------------------
    model = cp_model.CpModel()
    x: dict[tuple[int, int], cp_model.IntVar] = {}
    y: dict[tuple[int, int], cp_model.IntVar] = {}
    deviations: list[cp_model.IntVar] = []

    for i in range(sku_count):
        qty = quantities[i]
        for p in range(pallet_count):
            x[i, p] = model.NewIntVar(0, qty, f"x_{i}_{p}")
            y[i, p] = model.NewBoolVar(f"y_{i}_{p}")
            model.Add(x[i, p] <= qty * y[i, p])
            model.Add(x[i, p] >= y[i, p])

            dev = model.NewIntVar(0, pallet_count * qty, f"dev_{i}_{p}")
            model.AddAbsEquality(dev, pallet_count * x[i, p] - qty)
            deviations.append(dev)

        model.Add(sum(x[i, p] for p in range(pallet_count)) == qty)

    pallet_volume_vars: list[cp_model.IntVar] = []
    for p in range(pallet_count):
        model.Add(sum(y[i, p] for i in range(sku_count)) >= 2)
        volume_var = model.NewIntVar(0, capacity_units, f"vol_{p}")
        model.Add(
            volume_var
            == sum(volume_units[i] * x[i, p] for i in range(sku_count))
        )
        model.Add(volume_var <= capacity_units)
        pallet_volume_vars.append(volume_var)

    for p in range(pallet_count - 1):
        model.Add(pallet_volume_vars[p] >= pallet_volume_vars[p + 1])

    total_deviation = sum(deviations)
    model.Minimize(total_deviation)

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit_seconds
    solver.parameters.num_search_workers = 8
    solver.parameters.random_seed = 42

    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise AllocationError(
            "在固定托数、每托至少 2 个 SKU、且单托体积不超限的条件下，没有找到可行分配。"
        )

    phase1_solution = [
        [int(solver.Value(x[i, p])) for p in range(pallet_count)]
        for i in range(sku_count)
    ]
    best_deviation = int(round(solver.ObjectiveValue()))

    model.Add(total_deviation == best_deviation)
    max_volume = model.NewIntVar(0, capacity_units, "max_volume")
    min_volume = model.NewIntVar(0, capacity_units, "min_volume")
    for volume_var in pallet_volume_vars:
        model.Add(max_volume >= volume_var)
        model.Add(min_volume <= volume_var)

    model.ClearObjective()
    model.Minimize(max_volume - min_volume)

    solver2 = cp_model.CpSolver()
    solver2.parameters.max_time_in_seconds = time_limit_seconds
    solver2.parameters.num_search_workers = 8
    solver2.parameters.random_seed = 42

    status2 = solver2.Solve(model)
    if status2 not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return phase1_solution

    return [
        [int(solver2.Value(x[i, p])) for p in range(pallet_count)]
        for i in range(sku_count)
    ]

def allocate_mixed_pallets(group: pd.DataFrame, arn: str) -> tuple[list[dict[str, Any]], int]:
    pallet_values = pd.to_numeric(group[ORDER_PALLETS], errors="coerce").dropna()
    if pallet_values.empty:
        raise AllocationError("混合打托模式缺少 Pallets 总数。")

    invalid = pallet_values[(pallet_values <= 0) | ((pallet_values % 1).abs() > 1e-9)]
    if not invalid.empty:
        raise AllocationError("混合打托的 Pallets 必须是正整数。")

    unique_counts = sorted({int(v) for v in pallet_values})
    if len(unique_counts) != 1:
        raise AllocationError(
            f"同一个混合打托 ARN 出现多个不同的 Pallets 值：{unique_counts}。"
            "混合打托应填写一个固定的 ARN 总托数。"
        )
    pallet_count = unique_counts[0]

    # Aggregate duplicate rows of the same matched SKU.
    sku_rows: list[pd.Series] = []
    for _, sku_group in group.groupby("_product_sku_key", sort=False):
        first = sku_group.iloc[0].copy()
        first["_qty_num"] = int(sku_group["_qty_num"].sum())
        sku_rows.append(first)

    allocation = _solve_mixed_distribution(sku_rows, pallet_count)

    details: list[dict[str, Any]] = []
    for sku_index, row in enumerate(sku_rows):
        for pallet_index in range(pallet_count):
            qty = allocation[sku_index][pallet_index]
            if qty <= 0:
                continue
            details.append(
                _base_detail_record(
                    row=row,
                    arn=arn,
                    pallet_no=pallet_index + 1,
                    method_label="混合打托",
                    allocated_boxes=qty,
                )
            )

    # Final hard validation.
    detail_df = pd.DataFrame(details)
    if detail_df.empty:
        raise AllocationError("混合打托计算结果为空。")

    check = detail_df.groupby("Pallet No.").agg(
        SKU_Count=("UPC/SKU", "nunique"),
        Volume=("Allocated Volume (in³)", "sum"),
    )
    if len(check) != pallet_count:
        raise AllocationError("生成的托盘数量与固定 Pallets 总数不一致。")
    if (check["SKU_Count"] < 2).any():
        raise AllocationError("内部校验失败：存在少于 2 个不同 SKU 的混合托盘。")
    if (check["Volume"] > PALLET_CAPACITY_IN3 + 1e-6).any():
        raise AllocationError("内部校验失败：存在体积超过 48×40×72 in³ 的托盘。")

    return details, pallet_count


def allocate_orders(order_source: Any, product_source: Any) -> AllocationResult:
    order = load_order(order_source)
    product, product_warnings = load_product_master(product_source)
    order = enrich_order_with_product(order, product)

    all_details: list[dict[str, Any]] = []
    arn_summaries: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []

    for arn, group in order.groupby("_arn_key", sort=False):
        methods = unique_nonblank(group[ORDER_METHOD])
        po = " / ".join(unique_nonblank(group.get(ORDER_PO, pd.Series(dtype=object))))
        fc = " / ".join(unique_nonblank(group.get(ORDER_FC, pd.Series(dtype=object))))
        total_boxes = int(group["_qty_num"].sum())
        total_volume = (
            group["_qty_num"] * group["_carton_volume_in3"].fillna(0)
        ).sum()

        summary = {
            "Shipment ID (ARN)": arn,
            "Order/PO Number": po,
            "Fulfillment Center": fc,
            "打托方式": methods[0] if len(methods) == 1 else " / ".join(methods),
            "SKU Count": int(group["_product_sku_key"].replace("", pd.NA).nunique()),
            "Total Boxes": total_boxes,
            "Required / Generated Pallets": None,
            "Total Volume (in³)": float(total_volume),
            "Total Capacity (in³)": None,
            "Overall Utilization": None,
            "Status": "FAILED",
            "Message": "",
        }

        try:
            if len(methods) != 1:
                raise AllocationError(
                    f"同一个 ARN 必须有且只有一种打托方式。当前读取到：{methods or ['空白']}"
                )

            method_type = normalize_method(methods[0])
            if method_type == "unknown":
                raise AllocationError(f"无法识别打托方式：{methods[0]}")

            _validate_product_dimensions(group, arn)

            if method_type == "single":
                details, pallet_count = allocate_single_sku_pallets(group, arn)
            else:
                details, pallet_count = allocate_mixed_pallets(group, arn)

            # Verify allocated quantities equal order quantities by matched SKU.
            source_qty = (
                group.groupby("_product_sku_key", sort=False)["_qty_num"].sum().to_dict()
            )
            detail_df = pd.DataFrame(details)
            allocated_qty: dict[str, int] = {}
            for _, detail_row in detail_df.iterrows():
                key = normalize_code(detail_row["UPC/SKU"])
                # In normal flow UPC/SKU is the order UPC. Fallback by model is rare.
                if key not in source_qty:
                    model_key = normalize_model(detail_row["Model Number"])
                    matching_source = group[group["_model_key"] == model_key]
                    if not matching_source.empty:
                        key = matching_source.iloc[0]["_product_sku_key"]
                allocated_qty[key] = allocated_qty.get(key, 0) + int(detail_row["Allocated Boxes"])

            # Compare using product SKU keys to avoid formatting differences.
            for source_key, qty in source_qty.items():
                if allocated_qty.get(source_key, 0) != int(qty):
                    raise AllocationError(
                        f"内部数量校验失败：SKU {source_key} 订单 {qty} 箱，分配 {allocated_qty.get(source_key, 0)} 箱。"
                    )

            all_details.extend(details)
            capacity = pallet_count * PALLET_CAPACITY_IN3
            summary.update(
                {
                    "Required / Generated Pallets": pallet_count,
                    "Total Capacity (in³)": capacity,
                    "Overall Utilization": float(total_volume / capacity) if capacity else None,
                    "Status": "OK",
                    "Message": "分配完成",
                }
            )
        except Exception as exc:
            message = str(exc)
            summary["Message"] = message
            issues.append(
                {
                    "Shipment ID (ARN)": arn,
                    "Order/PO Number": po,
                    "Fulfillment Center": fc,
                    "打托方式": summary["打托方式"],
                    "错误说明": message,
                }
            )

        arn_summaries.append(summary)

    detail_columns = [
        "Shipment ID (ARN)",
        "Pallet No.",
        "打托方式",
        "Order/PO Number",
        "Fulfillment Center",
        "Model Number",
        "UPC/SKU",
        "产品匹配方式",
        "Allocated Boxes",
        "Weight per Box (lb)",
        "Length (in)",
        "Width (in)",
        "Height (in)",
        "Volume per Box (in³)",
        "Allocated Weight (lb)",
        "Allocated Volume (in³)",
    ]
    detail_df = pd.DataFrame(all_details, columns=detail_columns)

    if detail_df.empty:
        pallet_summary = pd.DataFrame(
            columns=[
                "Shipment ID (ARN)",
                "Pallet No.",
                "打托方式",
                "SKU Count",
                "Total Boxes",
                "Total Weight (lb)",
                "Total Volume (in³)",
                "Pallet Capacity (in³)",
                "Volume Utilization",
                "Status",
            ]
        )
    else:
        pallet_summary = (
            detail_df.groupby(
                ["Shipment ID (ARN)", "Pallet No.", "打托方式"],
                sort=False,
                as_index=False,
            )
            .agg(
                **{
                    "SKU Count": ("UPC/SKU", "nunique"),
                    "Total Boxes": ("Allocated Boxes", "sum"),
                    "Total Weight (lb)": ("Allocated Weight (lb)", "sum"),
                    "Total Volume (in³)": ("Allocated Volume (in³)", "sum"),
                }
            )
        )
        pallet_summary["Pallet Capacity (in³)"] = PALLET_CAPACITY_IN3
        pallet_summary["Volume Utilization"] = (
            pallet_summary["Total Volume (in³)"] / PALLET_CAPACITY_IN3
        )
        pallet_summary["Status"] = "OK"
        pallet_summary.loc[
            pallet_summary["Total Volume (in³)"] > PALLET_CAPACITY_IN3 + 1e-6,
            "Status",
        ] = "OVER VOLUME"
        mixed_bad = (
            pallet_summary["打托方式"].eq("混合打托")
            & (pallet_summary["SKU Count"] < 2)
        )
        pallet_summary.loc[mixed_bad, "Status"] = "LESS THAN 2 SKUs"

    arn_summary_df = pd.DataFrame(arn_summaries)
    issues_df = pd.DataFrame(
        issues,
        columns=[
            "Shipment ID (ARN)",
            "Order/PO Number",
            "Fulfillment Center",
            "打托方式",
            "错误说明",
        ],
    )

    return AllocationResult(
        detail=detail_df,
        pallet_summary=pallet_summary,
        arn_summary=arn_summary_df,
        issues=issues_df,
        product_warnings=product_warnings,
    )


def _style_worksheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells[:2000]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        width = min(max(max_length + 2, 10), 38)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def build_output_excel(result: AllocationResult) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.detail.to_excel(writer, sheet_name="托盘分配明细", index=False)
        result.pallet_summary.to_excel(writer, sheet_name="托盘汇总", index=False)
        result.arn_summary.to_excel(writer, sheet_name="ARN汇总", index=False)
        result.issues.to_excel(writer, sheet_name="错误与警告", index=False)
        result.product_warnings.to_excel(writer, sheet_name="产品资料警告", index=False)

        workbook = writer.book
        for ws in workbook.worksheets:
            _style_worksheet(ws)

        # Number formats
        for ws_name in ["托盘分配明细", "托盘汇总", "ARN汇总"]:
            ws = workbook[ws_name]
            headers = {cell.value: cell.column for cell in ws[1]}
            for header, col_idx in headers.items():
                if header and ("Volume" in str(header) or "Weight" in str(header)):
                    for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                        for c in cell:
                            c.number_format = "0.00"
                if header in {"Volume Utilization", "Overall Utilization"}:
                    for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                        for c in cell:
                            c.number_format = "0.0%"

    output.seek(0)
    return output.getvalue()


def load_fixed_product_file(path: str | Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    return load_product_master(path)
